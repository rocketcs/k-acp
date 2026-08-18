#!/usr/bin/env python3
"""Build a traceable Neo4j import dataset from the six Henan catalog XLSX files.

The source workbooks contain formatted empty tails, so this program streams rows
and accepts only rows with at least one value.  It deliberately keeps mapping
codes as MappingConcept evidence and never resolves them to ServiceItem.
"""
from __future__ import annotations

import argparse
import base64
import csv
import datetime as dt
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

ROOT = Path('/Users/rocket/kingsware/k-acp-query-mcp/data')
FILES = {
    'consumables_main': ROOT / '医疗服务项目及耗材映射库/2.河南省耗材映射库（2025年耗材目录）-20251013版+3月10日精神治疗等类整合项目.xlsx',
    'service_catalog': ROOT / '医疗服务项目及耗材映射库/1-河南服务项目目录（含映射库）-20260310（下发版）.xlsx',
    'basic_consumables': ROOT / '医疗服务项目及耗材映射库/2-2.河南省耗材映射库（2025年耗材目录）-基本物耗-20251013版+3月10日精神治疗等类整合项目.xlsx',
    'negotiated_consumables': ROOT / '医疗服务项目及耗材映射库/2-1.河南省耗材映射库（2025年耗材目录）-附件谈判子库-20251013版+3月10日精神治疗等类整合项目.xlsx',
    'drug_catalog': ROOT / '药品目录模板202603.xlsx',
    'diagnosis_catalog': ROOT / '诊疗目录模板202509.xlsx',
}
SERVICE_HEADERS = ['序号', '财务分类代码', '河南省项目编码', '河南省项目名称', '项目内涵', '除外内容', '服务产出', '价格构成', '加收项', '扩展项', '计价单位', '说明', '医保支付类别', '备注（医保支付）', '拼音助记码', '国家医疗服务项目代码', '国家医疗服务项目名称', '开始时间', '终止时间', '经办人', '经办时间', '职工首付比例', '居民首付比例', '离休首付比例', '生育首付比例', '离休最高限额', '省级二档最高限额', '省级一档最高限额', '市级三档最高限额', '市级二档最高限额', '市级一档最高限额', '县区三级最高限额', '县区二级最高限额', '县区一级最高限额', '乡级最高限额', '政策号']
SHEETS = [
    ('consumables_main', '耗材映射库', 2, 'consumable'), ('consumables_main', '变更', 2, 'consumable_change'),
    ('service_catalog', '服务项目映射库', 5, 'service'), ('service_catalog', '新增', 5, 'service_added'), ('service_catalog', '停用', 5, 'service_stopped'),
    ('basic_consumables', '基本物耗映射库', 2, 'basic_material'), ('negotiated_consumables', '映射库谈判子库', 2, 'consumable_sku'),
    ('drug_catalog', '药品目录模板', 2, 'drug'),
    ('diagnosis_catalog', '诊疗目录模板', 2, 'diagnosis'), ('diagnosis_catalog', 'Sheet1', 1, 'diagnosis_supplement'), ('diagnosis_catalog', '二级代码对照', 1, 'lookup'),
]
TOKEN_SPLIT = re.compile(r'[,，;；\n]+')
TOKEN_CODE = re.compile(r'^([A-Za-z0-9]+)')
PIPELINE_VERSION = 'henan-neo4j-import-v1.0.0'

def sha(value: str) -> str: return hashlib.sha256(value.encode('utf-8')).hexdigest()
def clean(value):
    if value is None: return ''
    if isinstance(value, (dt.date, dt.datetime)): return value.isoformat()[:10]
    return unicodedata.normalize('NFKC', str(value)).replace('\xa0', ' ').strip()
def date_value(value):
    value = clean(value)
    if not value: return ''
    try:
        if re.fullmatch(r'\d+(\.\d+)?', value): return from_excel(float(value)).date().isoformat()
        return dt.datetime.strptime(value.replace('/', '-').replace('.', '-'), '%Y-%m-%d').date().isoformat()
    except (ValueError, TypeError, OverflowError): return ''
def decimal_value(value):
    text = clean(value).replace(',', '')
    try: return str(float(text)) if text else ''
    except ValueError: return ''
def ident(*values): return sha('|'.join(clean(v) for v in values))
def prefix(value):
    value = clean(value)
    return value.split('-', 1) if '-' in value else (value, value)
def semantics(value):
    number = decimal_value(value)
    return 'UNCONFIRMED_SENTINEL' if number in {'20000.0', '40000.0', '110000.0'} else ('OBSERVED_VALUE' if number else '')

class Dataset:
    def __init__(self, output: Path):
        self.output, self.rows = output, defaultdict(list)
        self.metrics = {'pipeline_version': PIPELINE_VERSION, 'created_at': dt.datetime.now().isoformat(), 'sheets': {}, 'gates': [], 'warnings': []}
        self.file_hash = {key: hashlib.sha256(path.read_bytes()).hexdigest() for key, path in FILES.items()}
        manifest = '\n'.join(f'{key}|{self.file_hash[key]}' for key in sorted(self.file_hash))
        self.batch_id = sha(manifest + '|' + PIPELINE_VERSION + '|full')
        self.main_codes, self.basic_codes, self.change_codes, self.sku_codes = set(), set(), set(), set()
        self.mapping_tokens, self.mapping_unparsed, self.mapping_duplicates = 0, 0, 0
        self.service_codes, self.national_service_codes, self.mapping_codes = set(), set(), set()
        self.business_rows = 0
        self.consumable_categories, self.mapping_concepts, self.mapping_labels = {}, {}, {}
        self.bases, self.products, self.skus, self.basic_materials = {}, {}, {}, {}
        self.service_categories, self.services, self.diagnoses, self.drugs = {}, {}, {}, {}

    def add(self, name, row): self.rows[name].append(row)
    def read(self, file_key, sheet, start, kind):
        path = FILES[file_key]; wb = load_workbook(path, read_only=True, data_only=True); ws = wb[sheet]
        if kind.startswith('service'): headers = SERVICE_HEADERS
        elif sheet == 'Sheet1':
            dwb = load_workbook(FILES['diagnosis_catalog'], read_only=True, data_only=True)
            headers = [clean(v) or f'column_{i+1}' for i, v in enumerate(next(dwb['诊疗目录模板'].iter_rows(min_row=1, max_row=1, values_only=True)))]
            dwb.close()
        else: headers = [clean(v) or f'column_{i+1}' for i, v in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))]
        source_id = sha(self.file_hash[file_key] + '|' + sheet)
        self.add('source_files', {'source_id': source_id, 'batch_id': self.batch_id, 'file_name': path.name, 'sheet_name': sheet, 'file_sha256': self.file_hash[file_key]})
        accepted = 0
        for excel_row, values in enumerate(ws.iter_rows(min_row=start, values_only=True), start=start):
            values = [clean(v) for v in values]
            if not any(values): continue
            record = {headers[i] if i < len(headers) else f'column_{i+1}': values[i] for i in range(len(values))}
            raw = json.dumps(record, ensure_ascii=False, sort_keys=True, separators=(',', ':'))
            record_id = sha(source_id + '|' + str(excel_row) + '|' + sha(raw))
            self.add('catalog_records', {'record_id': record_id, 'source_id': source_id, 'source_row': excel_row, 'raw_row_hash': sha(raw), 'raw_payload_base64': base64.b64encode(raw.encode('utf-8')).decode('ascii'), 'raw_payload_encoding': 'base64:utf-8-json', 'record_kind': kind})
            if kind != 'lookup': self.business_rows += 1
            accepted += 1
            self.domain(record, record_id, source_id, path.name, sheet, excel_row, kind)
        wb.close(); self.metrics['sheets'][f'{file_key}/{sheet}'] = {'accepted_rows': accepted, 'source_id': source_id, 'start_row': start}

    def category(self, r):
        parts = [('L1', r.get('一级分类','')), ('L2', r.get('二级分类','')), ('L3', r.get('三级分类','')), ('G', r.get('医保通用名分类','')), ('M', r.get('材质','')), ('F', r.get('特征',''))]
        parent = ''; final = ''
        for level, raw in parts:
            if not raw: continue
            code, name = prefix(raw); path = (parent + '/' if parent else '') + f'{level}:{code}'
            self.consumable_categories[path] = {'category_path_id': path, 'parent_path_id': parent, 'level': len(path.split('/')), 'code': code, 'name': name, 'path_kind': level}
            parent, final = path, path
        return final

    def mapping(self, r, record_id, source_id, file_name, sheet, excel_row, product_id='', relation='ASSERTED_MAPS_TO_CONCEPT', basic_material_id=''):
        field = '映射结果' if '映射结果' in r else '备注'
        raw_value = r.get(field, '')
        if not raw_value: return
        tokens = [t.strip() for t in TOKEN_SPLIT.split(raw_value) if t.strip()]
        self.mapping_tokens += len(tokens); self.mapping_duplicates += len(tokens) - len(set(tokens))
        for token in tokens:
            match = TOKEN_CODE.match(token)
            if not match: self.mapping_unparsed += 1; continue
            code = match.group(1); label = token[len(code):].strip(); namespace = 'HENAN_MAPPING_SOURCE'
            self.mapping_codes.add(code)
            self.mapping_concepts[(namespace, code)] = {'namespace': namespace, 'code': code, 'classification_status': 'PROJECT_LIKE' if len(code) >= 8 else 'UNKNOWN', 'classification_rule': 'alphanumeric_prefix_from_source_token', 'review_status': 'UNRESOLVED', 'first_seen_batch': self.batch_id}
            assertion_id = sha(record_id + '|' + field + '|' + token)
            if label:
                label_id = sha(namespace + '|' + code + '|' + label)
                self.mapping_labels[(label_id, assertion_id)] = {'observed_label_id': label_id, 'assertion_id': assertion_id, 'namespace': namespace, 'code': code, 'label': label, 'record_id': record_id, 'source_row': excel_row}
            if product_id:
                self.add('mapping_assertions', {'assertion_id': assertion_id, 'record_id': record_id, 'product_id': product_id, 'namespace': namespace, 'code': code, 'batch_id': self.batch_id, 'source_file': file_name, 'source_sheet': sheet, 'source_row': excel_row, 'source_column': field, 'raw_token': token, 'valid_from': date_value(r.get('开始时间','')), 'valid_to': date_value(r.get('终止时间','')), 'relation_type': relation, 'basic_material_id': basic_material_id})
            else:
                self.add('concept_mentions', {'assertion_id': assertion_id, 'record_id': record_id, 'namespace': namespace, 'code': code, 'source_row': excel_row, 'raw_token': token})

    def domain(self, r, record_id, source_id, file_name, sheet, excel_row, kind):
        if kind.startswith('consumable') or kind == 'basic_material':
            code, registration = r.get('耗材代码',''), r.get('注册备案号','')
            if not code: return
            base_code = code[:20] if len(code) == 27 else code
            if kind == 'consumable': self.main_codes.add(code)
            elif kind == 'consumable_change': self.change_codes.add(code)
            elif kind == 'basic_material': self.basic_codes.add(code)
            else: self.sku_codes.add(code)
            path = self.category(r)
            self.bases.setdefault(base_code, {'base_code': base_code, 'record_id': record_id, 'feature_path_id': path, 'code_rule_valid': str(len(base_code) == 20).lower()})
            product_id = ident(base_code, registration)
            self.products.setdefault(product_id, {'product_id': product_id, 'base_code': base_code, 'registration_no': registration, 'registration_id': ident('registration', registration) if registration else '', 'product_name': r.get('单件产品名称','')})
            self.add('consumable_product_evidence', {'record_id': record_id, 'product_id': product_id, 'base_code': base_code})
            relation, material_id = 'ASSERTED_MAPS_TO_CONCEPT', ''
            if kind == 'basic_material':
                material = r.get('基本物耗','')
                material_id = ident('basic_material', material) if material else ''
                if material_id: self.basic_materials.setdefault(material_id, {'basic_material_id': material_id, 'name': material})
                relation = 'ASSERTED_BASIC_FOR_CONCEPT'
            self.mapping(r, record_id, source_id, file_name, sheet, excel_row, product_id, relation, material_id)
            if kind == 'consumable_sku': self.skus.setdefault(code, {'sku_code': code, 'base_code': base_code, 'product_id': product_id if registration else '', 'record_id': record_id, 'specification': r.get('规格',''), 'model': r.get('型号','')})
        elif kind.startswith('service'):
            code, name = r.get('河南省项目编码',''), r.get('河南省项目名称','')
            if not code: return
            self.service_codes.add(code)
            national_code = r.get('国家医疗服务项目代码','')
            if national_code: self.national_service_codes.add(national_code)
            valid_from, valid_to = date_value(r.get('开始时间','')), date_value(r.get('终止时间',''))
            if len(code) == 15:
                sid = 'HENAN_SERVICE|' + code
                self.services.setdefault(sid, {'service_item_id': sid, 'province_code': code, 'name': name, 'record_id': record_id, 'version_id': ident('SERVICE', sid, record_id), 'record_hash': sha(json.dumps(r, ensure_ascii=False, sort_keys=True)), 'valid_from': valid_from, 'valid_to': valid_to, 'raw_valid_from': r.get('开始时间',''), 'raw_valid_to': r.get('终止时间','')})
            else:
                cid = 'HENAN_SERVICE_CATEGORY|' + code
                self.service_categories.setdefault(cid, {'service_category_id': cid, 'namespace': 'HENAN_SERVICE', 'code': code, 'name': name, 'parent_id': ''})
        elif kind.startswith('diagnosis'):
            code, name = r.get('项目编码',''), r.get('项目名称','')
            if code:
                did = 'HENAN_DIAGNOSIS|' + code
                self.diagnoses.setdefault(did, {'diagnosis_item_id': did, 'namespace': 'HENAN_DIAGNOSIS', 'code': code, 'name': name, 'record_id': record_id, 'valid_from': date_value(r.get('开始时间','')), 'valid_to': date_value(r.get('终止时间',''))})
            if kind == 'diagnosis_supplement': self.mapping(r, record_id, source_id, file_name, sheet, excel_row)
        elif kind == 'drug':
            code = r.get('药品编码','')
            if code:
                generic = r.get('通用名称',''); org = r.get('药厂名称',''); gid = ident('drug_generic', generic)
                self.drugs.setdefault(code, {'drug_code': code, 'record_id': record_id, 'drug_generic_id': gid, 'generic_name': generic, 'specification': r.get('规格',''), 'max_price_raw': r.get('最高价格',''), 'max_price': decimal_value(r.get('最高价格','')), 'max_price_semantics': semantics(r.get('最高价格','')), 'organization_id': ident('organization', org) if org else '', 'organization_alias_id': ident('organization_alias', org) if org else '', 'organization_normalized_name': org, 'organization_raw_name': org})

    def validate(self):
        service_aliases = set(self.service_codes) | set(self.national_service_codes)
        for code in list(service_aliases):
            if '-' in code: service_aliases.update(part for part in code.split('-') if part)
        checks = [
            ('business_row_count', self.business_rows == 199570, self.business_rows, 199570),
            ('main_consumable_rows', self.metrics['sheets']['consumables_main/耗材映射库']['accepted_rows'] == 157738, self.metrics['sheets']['consumables_main/耗材映射库']['accepted_rows'], 157738),
            ('base_code_length', all(len(code) == 20 for code in self.main_codes), len(self.main_codes), 'all 20 chars'),
            ('sku_code_length', all(len(code) == 27 for code in self.sku_codes), len(self.sku_codes), 'all 27 chars'),
            ('basic_coverage', self.basic_codes <= self.main_codes, len(self.basic_codes & self.main_codes), len(self.basic_codes)),
            ('change_coverage', self.change_codes <= self.main_codes, len(self.change_codes & self.main_codes), len(self.change_codes)),
            ('sku_prefix_coverage', all(code[:20] in self.main_codes for code in self.sku_codes), sum(code[:20] in self.main_codes for code in self.sku_codes), len(self.sku_codes)),
            ('mapping_tokens', self.mapping_tokens == 665401, self.mapping_tokens, 665401),
            ('mapping_concept_count', len(self.mapping_codes) == 601, len(self.mapping_codes), 601),
            ('mapping_service_alias_overlap', len(self.mapping_codes & service_aliases) == 1, len(self.mapping_codes & service_aliases), 1),
            ('mapping_unparsed', self.mapping_unparsed == 0, self.mapping_unparsed, 0),
            ('mapping_duplicate_tokens', self.mapping_duplicates == 0, self.mapping_duplicates, 0),
        ]
        self.metrics['gates'] = [{'name': n, 'passed': ok, 'actual': actual, 'expected': expected} for n, ok, actual, expected in checks]
        self.metrics['mapping_service_alias_overlap_codes'] = sorted(self.mapping_codes & service_aliases)
        self.metrics['passed'] = all(item['passed'] for item in self.metrics['gates'])

    def write_csv(self):
        self.rows['import_batches'] = [{'batch_id': self.batch_id, 'input_manifest_sha256': sha('\n'.join(sorted(self.file_hash.values()))), 'pipeline_version': PIPELINE_VERSION, 'run_mode': 'full', 'started_at': self.metrics['created_at']}]
        for name, values in [('consumable_categories', self.consumable_categories.values()), ('consumable_bases', self.bases.values()), ('consumable_products', self.products.values()), ('consumable_skus', self.skus.values()), ('basic_materials', self.basic_materials.values()), ('mapping_concepts', self.mapping_concepts.values()), ('mapping_labels', self.mapping_labels.values()), ('service_categories', self.service_categories.values()), ('service_items', self.services.values()), ('diagnosis_items', self.diagnoses.values()), ('drug_products', self.drugs.values())]: self.rows[name] = list(values)
        for name, values in self.rows.items():
            if not values: continue
            fields = list(dict.fromkeys(k for row in values for k in row))
            with (self.output / f'{name}.csv').open('w', encoding='utf-8', newline='') as f:
                w = csv.DictWriter(f, fieldnames=fields, extrasaction='ignore'); w.writeheader(); w.writerows(values)
        (self.output / 'validation_report.json').write_text(json.dumps(self.metrics, ensure_ascii=False, indent=2), encoding='utf-8')

    def run(self):
        self.output.mkdir(parents=True, exist_ok=True)
        for spec in SHEETS: self.read(*spec)
        self.validate(); self.write_csv()
        print(json.dumps({'output': str(self.output), 'batch_id': self.batch_id, 'passed': self.metrics['passed'], 'business_rows': self.business_rows, 'mapping_tokens': self.mapping_tokens}, ensure_ascii=False))
        return self.metrics['passed']

if __name__ == '__main__':
    ap = argparse.ArgumentParser(); ap.add_argument('--output', type=Path, required=True); args = ap.parse_args()
    if args.output.exists() and any(args.output.iterdir()):
        raise SystemExit(f'Output directory must be new or empty: {args.output}')
    raise SystemExit(0 if Dataset(args.output).run() else 2)
